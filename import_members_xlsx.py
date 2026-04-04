#!/usr/bin/env python3
"""Import des membres depuis le fichier Excel vue_membre_users.
Efface tous les comptes existants puis importe les 314 membres.
Usage: DATABASE_URL='postgresql://...' python import_members_xlsx.py
"""
import os, sys

DB_URL = os.environ.get('DATABASE_URL')
if not DB_URL:
    print("ERREUR: DATABASE_URL requis.")
    print("Usage: DATABASE_URL='postgresql://...' python import_members_xlsx.py")
    sys.exit(1)

import openpyxl
import psycopg2
import psycopg2.extras
import bcrypt

XLSX_FILE = os.path.join(os.path.dirname(__file__), 'vue_membre_users 2026_03_16.xlsx')
DEFAULT_PASSWORD = 'Affi2026!'  # Sera hashé en bcrypt

# === Mapping des types de cotisation vers membership_type ===
TYPE_MAP = {
    'Ingenieur -Cadre de plus de 30 ans': 'standard',
    'Ingenieur-Cadre < \xe0 30ans': 'jeune',
    'Retrait\xe9': 'retraite',
    'Retraité': 'retraite',
    'Etudiant': 'etudiant',
    'Exempté': 'honneur',
    'Exempt\xe9': 'honneur',
    'Gratuit 1 an': 'standard',
}

# === Mapping des statuts vers status ===
STATUS_MAP = {
    'A jour': 'active',
    'Rappel': 'active',       # rappel de cotisation mais toujours membre
    '2 ans': 'active',        # 2 ans d'ancienneté
    'Pr\xe9radiation': 'pending',  # pré-radiation = en sursis
    'Préradiation': 'pending',
    'Radiation': 'blocked',    # radié = bloqué
    'NULL': 'pending',
}

def hash_password(pw):
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()

def clean_name(name):
    """Nettoie et formate un nom/prénom."""
    if not name:
        return ''
    name = str(name).strip()
    # Capitalise correctement (ROGER -> Roger, JEAN-LOUIS -> Jean-Louis)
    parts = name.split('-')
    return '-'.join(p.capitalize() for p in parts)

def main():
    print(f"Lecture du fichier: {XLSX_FILE}")
    wb = openpyxl.load_workbook(XLSX_FILE)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    # Lire toutes les lignes
    members = []
    for row_idx in range(2, ws.max_row + 1):
        row = {}
        for col_idx, header in enumerate(headers):
            row[str(header)] = ws.cell(row=row_idx, column=col_idx + 1).value
        members.append(row)

    print(f"Membres lus: {len(members)}")

    # Connexion DB
    conn = psycopg2.connect(DB_URL, sslmode='require')
    cur = conn.cursor()

    # === ETAPE 1: Effacer tous les comptes existants ===
    print("\n=== SUPPRESSION DES COMPTES EXISTANTS ===")
    # Ordre important: respecter les foreign keys
    tables_to_clean = [
        'feed_likes', 'feed_comments', 'feed_posts',
        'poll_votes', 'poll_options', 'polls',
        'notifications', 'messages', 'endorsements',
        'job_applications', 'jobs', 'member_projects',
        'member_announcements', 'event_communications',
        'event_registrations', 'quizz_scores',
        'sessions', 'consentements',
        'subscriptions', 'board_members',
        'newsletter_subscribers', 'logs',
    ]
    for table in tables_to_clean:
        try:
            cur.execute(f"DELETE FROM {table}")
            print(f"  {table}: videe")
        except Exception as e:
            print(f"  {table}: skip ({e})")
            conn.rollback()
            continue

    # Supprimer tous les membres (sauf contraintes)
    cur.execute("DELETE FROM members")
    deleted = cur.rowcount
    print(f"  members: {deleted} supprime(s)")
    conn.commit()

    # Reset la sequence auto-increment
    cur.execute("ALTER SEQUENCE members_id_seq RESTART WITH 1")
    conn.commit()

    # === ETAPE 2: Hasher le mot de passe par defaut ===
    print(f"\nHashage du mot de passe par defaut...")
    pw_hash = hash_password(DEFAULT_PASSWORD)

    # === ETAPE 3: Importer les membres ===
    print(f"\n=== IMPORT DES {len(members)} MEMBRES ===")
    imported = 0
    skipped = 0
    errors = []

    for i, m in enumerate(members):
        nom = clean_name(m.get('nom', ''))
        prenom = clean_name(m.get('prenom', ''))
        email = (m.get('email') or '').strip().lower()
        statut_src = str(m.get('statutMembre') or 'NULL')
        type_src = str(m.get('typeCotisation') or '')
        code = m.get('code', '')

        # Skip si pas d'email
        if not email or '@' not in email:
            skipped += 1
            errors.append(f"Ligne {i+2}: pas d'email ({prenom} {nom})")
            continue

        # Mapping
        status = STATUS_MAP.get(statut_src, 'pending')
        membership_type = TYPE_MAP.get(type_src, 'standard')

        try:
            cur.execute("""
                INSERT INTO members (
                    email, password_hash, first_name, last_name,
                    membership_type, status, role,
                    consent_annuaire, consent_newsletter, consent_date,
                    joined_at, created_at
                ) VALUES (
                    %s, %s, %s, %s,
                    %s, %s, 'member',
                    FALSE, FALSE, NOW(),
                    NOW(), NOW()
                ) ON CONFLICT (email) DO NOTHING
            """, [
                email, pw_hash, prenom, nom,
                membership_type, status,
            ])
            if cur.rowcount > 0:
                imported += 1
            else:
                skipped += 1
                errors.append(f"Ligne {i+2}: email duplique ({email})")
        except Exception as e:
            errors.append(f"Ligne {i+2}: {str(e)[:80]}")
            conn.rollback()

    conn.commit()

    # === ETAPE 4: Creer le compte admin ===
    print("\n=== CREATION COMPTE ADMIN ===")
    admin_hash = hash_password('Admin2026!')
    cur.execute("""
        INSERT INTO members (
            email, password_hash, first_name, last_name,
            membership_type, status, role, is_admin, is_board,
            consent_annuaire, consent_newsletter, consent_date
        ) VALUES (
            'admin@ingenieur-ferroviaire.net', %s, 'Admin', 'AFFI',
            'standard', 'active', 'admin', TRUE, TRUE,
            TRUE, TRUE, NOW()
        ) ON CONFLICT (email) DO UPDATE SET
            password_hash = EXCLUDED.password_hash,
            is_admin = TRUE, status = 'active'
    """, [admin_hash])
    conn.commit()
    print("  Admin cree: admin@ingenieur-ferroviaire.net")

    # === RESUME ===
    cur.execute("SELECT COUNT(*) FROM members")
    total = cur.fetchone()[0]
    cur.execute("SELECT status, COUNT(*) FROM members GROUP BY status ORDER BY COUNT(*) DESC")
    by_status = cur.fetchall()
    cur.execute("SELECT membership_type, COUNT(*) FROM members GROUP BY membership_type ORDER BY COUNT(*) DESC")
    by_type = cur.fetchall()

    print(f"\n{'='*50}")
    print(f"IMPORT TERMINE")
    print(f"{'='*50}")
    print(f"  Importes: {imported}")
    print(f"  Ignores:  {skipped}")
    print(f"  Total en base: {total}")
    print(f"\n  Par statut:")
    for s, n in by_status:
        print(f"    {s}: {n}")
    print(f"\n  Par type:")
    for t, n in by_type:
        print(f"    {t}: {n}")

    if errors:
        print(f"\n  Erreurs ({len(errors)}):")
        for e in errors[:10]:
            print(f"    {e}")
        if len(errors) > 10:
            print(f"    ... et {len(errors)-10} autres")

    conn.close()
    print(f"\nMot de passe par defaut: {DEFAULT_PASSWORD}")
    print("Admin: admin@ingenieur-ferroviaire.net / Admin2026!")

if __name__ == '__main__':
    main()
